import enum
from openpyxl import load_workbook, workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import threading
import os
import openpyxl
import concurrent.futures
import time
import datetime

datetime_now = datetime.datetime.now()

flashscore_main_site_url = 'https://www.flashscore.pl/'

excel_template = "/template.xlsx"
output_excel_name = ""
wb = openpyxl.Workbook()
ws = wb.active

home_name = []
away_name = []
link_to_details = []
saving_data = []

wb = load_workbook(filename="template.xlsx")
ws = wb.active

universal_class = "_simpleText_lsjrv_4"

chromedriver_path = os.path.join(os.getcwd(), 'chromedriver/chromedriver.exe')
service = Service(chromedriver_path)

# Set Chrome options for headless mode
chrome_options = Options()
chrome_options.add_argument("--disable-search-engine-choice-screen")
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu')

driver = webdriver.Chrome(service=service, options=chrome_options)
# driver.maximize_window()
driver.get(flashscore_main_site_url)
accept_cookies = driver.find_element(By.ID, "onetrust-accept-btn-handler").click()



def view_previous_day():
    print("view prev day")
    previous_day_button = driver.find_element(By.CLASS_NAME, "calendar__navigation--yesterday").click()

def reveal_all_events():
    driver.execute_script("""
    document.querySelectorAll('._simpleText_lsjrv_4').forEach((span_puppet, index) => {
        if (span_puppet.parentElement.nodeName === "BUTTON") {
            span_puppet.parentElement.click();
        };
    });
    """)

def scrape_all_events():
    home_name_scraped = driver.execute_script("""
        let event_divs = document.querySelectorAll(".event__match");
        let home_name_result = [];
        event_divs.forEach((event_div, index) => 
        {
        let home_name = event_div.querySelector(".event__homeParticipant").querySelector("._simpleText_lsjrv_4").textContent; 
        home_name_result.push(home_name)
        });     
        return home_name_result;
    """)
    
    home_name.extend(home_name_scraped)
    
    away_name_scraped = driver.execute_script("""
        let event_divs = document.querySelectorAll(".event__match");
        let away_name_result = [];
        event_divs.forEach((event_div, index) => 
        {
        let away_name = event_div.querySelector(".event__awayParticipant").querySelector("._simpleText_lsjrv_4").textContent; 
        away_name_result.push(away_name)
        });     
        return away_name_result;
    """)
    
    away_name.extend(away_name_scraped)
    
    link_to_details_scraped = driver.execute_script("""
        let event_divs = document.querySelectorAll(".event__match");
        let link_to_details_result = [];
        event_divs.forEach((event_div, index) => 
        {
        let link_to_details = event_div.querySelector(".eventRowLink").getAttribute("href").replace("#/szczegoly-meczu", "#/h2h/overall");
        link_to_details_result.push(link_to_details);
        });
        return link_to_details_result;
    """)
    
    link_to_details.extend(link_to_details_scraped)

    saving_data_scraped = driver.execute_script("""
        let event_divs = document.querySelectorAll(".event__match");
        let saving_data_result = [];
        event_divs.forEach((event_div, index) => 
        {
        let saving_data = document.getElementById("calendarMenu").textContent;
        saving_data_result.push(saving_data)
        });
        return saving_data_result;
    """)
    
    saving_data.extend(saving_data_scraped)
    
    print(f"HOME NAME: {home_name}")
    print(f"AWAY NAME: {away_name}")
    print(f"LINK DETAILS: {link_to_details}")
    print(f"SAVING DATA: {saving_data}")

def write_first_part_excel():
    print("First part to excel")
    for i, (home, away, link, date_date) in enumerate(zip(home_name, away_name, link_to_details, saving_data), start=2):
        print(i)
        ws.cell(row=i, column=3, value=home)
        ws.cell(row=i, column=4, value=away)
        ws.cell(row=i, column=2, value=link)
        ws.cell(row=i, column=1, value=date_date)

# Create a lock
write_lock = threading.Lock()

def scrape_from_link(process_number, equal_part, len_proc_num, driver):
    # Calculate start and end points for each driver
    where_to_start = process_number * equal_part
    where_to_end = where_to_start + equal_part if process_number != len_proc_num - 1 else len(link_to_details)
    cookie_clicked = False  # Track cookie acceptance for each driver

    for k, i in enumerate(link_to_details[where_to_start:where_to_end], start=where_to_start):
        print(f"{k} indexed link is beeing scraped")

        def the_rest(driver):
            tie_home = 0
            tie_away = 0
            tie_face = 0

            # Wait until the h2h section is loaded
            elem = WebDriverWait(driver, 5, poll_frequency=0.1).until(EC.presence_of_element_located((By.CLASS_NAME, "h2h__section")))

            # SECTION 1: Home team last matches
            event_section_one_events_icon = driver.find_elements(By.CLASS_NAME, "h2h__section")[0].find_elements(By.CLASS_NAME, "h2h__icon")
            event_section_one_events_outcome = []
    
            for x in range(len(event_section_one_events_icon)):
                event_section_one_events_outcome.append(event_section_one_events_icon[x].find_element(By.TAG_NAME, "div").get_attribute('title'))

            # Lock to ensure only one thread writes at a time
            with write_lock:
                for j, one_data in enumerate(event_section_one_events_outcome, start=5):
                    ws.cell(row=k+2, column=j, value=one_data)

                    # Count draws for the home team
                    if one_data == "Remis":
                        tie_home += 1

                ws.cell(row=k+2, column=10, value=tie_home)

                # SECTION 2: Away team last matches
                event_section_two = driver.find_elements(By.CLASS_NAME, "h2h__section")[1].find_elements(By.CLASS_NAME, "h2h__icon")
                event_section_two_events_outcome = []

                for x in range(len(event_section_two)):
                    event_section_two_events_outcome.append(event_section_two[x].find_element(By.TAG_NAME, "div").get_attribute('title'))

                # Write data to Excel for away team matches
                for j, two_data in enumerate(event_section_two_events_outcome, start=11):
                    ws.cell(row=k+2, column=j, value=two_data)

                    # Count draws for the away team
                    if two_data == "Remis":
                        tie_away += 1

                ws.cell(row=k+2, column=16, value=tie_away)

                # SECTION 3: Face-to-face results
                event_section_three_rows = driver.find_elements(By.CLASS_NAME, "h2h__section")[2].find_elements(By.CLASS_NAME, "h2h__row")
                event_section_three_events_outcome = []

                for x in range(len(event_section_three_rows)):
                    event_section_three_result = event_section_three_rows[x].find_element(By.CLASS_NAME, "h2h__result")
                    
                    try:
                        event_section_three_result_score_win = event_section_three_result.find_elements(By.TAG_NAME, "span")[0].get_attribute('innerHTML')
                        event_section_three_result_score_lose = event_section_three_result.find_elements(By.TAG_NAME, "span")[1].get_attribute('innerHTML')
                    except:
                        event_section_three_result_score_win = 69
                        event_section_three_result_score_lose = 420

                    event_section_three_events_outcome.append(f"{event_section_three_result_score_win} : {event_section_three_result_score_lose}")

                # Write data to Excel for face-to-face matches
                for j, three_data in enumerate(event_section_three_events_outcome, start=17):
                    ws.cell(row=k+2, column=j, value=three_data)

                    first_score, second_score = three_data.split(": ")

                    # Count draws in face-to-face matches
                    if int(first_score) == int(second_score):
                        tie_face += 1

                ws.cell(row=k+2, column=22, value=tie_face)

                # If there are at least 3 draws in all sections, mark the row
                if tie_home >= 3 and tie_away >= 3 and tie_face >= 3:
                    ws.cell(row=k+2, column=23, value=f"{tie_home}/{tie_away}/{tie_face}")

        
        try:
            driver.get(i)
        except:
            try:
                driver.get(i)
            except:
                try:
                    driver.get(i)
                except:
                    try:
                        driver.get(i)
                    except:
                        continue
                        

        if not cookie_clicked:
            try:
                accept_cookies = driver.find_element(By.ID, "onetrust-accept-btn-handler").click()
                cookie_clicked = True
            except Exception as e:
                print(f"Cookies already accepted or not found for process {process_number}: {str(e)}")

        the_rest(driver)

def run_in_parallel():
    print("Running in parallel...")
    process_numbers = [0, 1]  # Adjust as needed
    len_proc_num = len(process_numbers)
    total_links = len(link_to_details)

    equal_part = total_links // len(process_numbers)
    extra = total_links % len(process_numbers)

    # Initialize drivers for each thread
    drivers = []
    for _ in process_numbers:
        drivers.append(webdriver.Chrome(service=service, options=chrome_options))

    # Submit tasks to ThreadPoolExecutor
    with concurrent.futures.ThreadPoolExecutor(max_workers=len_proc_num) as executor:
        futures = [
            executor.submit(
                scrape_from_link, process_number, equal_part + (1 if process_number < extra else 0), len_proc_num, drivers[process_number]
            ) for process_number in process_numbers
        ]
        
        for future in concurrent.futures.as_completed(futures):
            try:
                result = future.result()
                print(result)
            except Exception as e:
                print(f"Error occurred in a thread: {str(e)}")

def main():
    
    for x in range(5):
        print(f"Days started {x}")
        reveal_all_events()
        scrape_all_events()
        view_previous_day()
        print(f"Scraped day {x}")
    
    write_first_part_excel()
    
    driver.quit()

    run_in_parallel()
    
    print("Saving excel")
    wb.save(f"data {datetime_now.strftime('%d')}-{datetime_now.strftime('%m')}-{datetime_now.strftime('%Y')} {datetime_now.strftime('%H')}-{datetime_now.strftime('%M')}.xlsx")
    
    wb.close()

main()